# masters/views.py
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import status
from django.db import transaction
from django.db.models import Sum, Prefetch, Q
from rest_framework.parsers import MultiPartParser, FormParser,JSONParser
from rest_framework import generics
from django.db.utils import IntegrityError
from rest_framework.pagination import PageNumberPagination
from .models import MasterlistDocument,Supplier, Grade, Customer, TypeOfMaterial,Location, RMReceiving, HoldMaterial, BatchTracking, TagGeneration
from forging.models import Forging
from .serializers import (
    SupplierSerializer,
    GradeSerializer,
    CustomerSerializer,
    TypeOfMaterialSerializer,
    LocationSerializer,RMReceivingSerializer, HoldMaterialSerializer, BatchTrackingSerializer,HoldMaterialListSerializer

)
from django.db.models.fields.files import FieldFile
from django.db.models import Case, When, F, Value, IntegerField

from rest_framework import serializers
from django.http import JsonResponse
from django.http import HttpResponse
import datetime
import openpyxl
from django.utils.dateparse import parse_date
from decimal import Decimal
from django.db import models
import uuid
from django.conf import settings
from machining.models import machining
from fi.models import Fi
from heat_treatment.models import HeatTreatment
from marking.models import marking
from pre_mc.models import pre_mc
from visual.models import Visual
from dispatch.models import dispatch
from django.db.models import Count, Min, Max
from .aggregates import GroupConcat  # Adjust the path if needed
from django.core.paginator import Paginator
from rest_framework.decorators import api_view, permission_classes
from .serializers import (
    MasterlistSerializer1,
    MasterlistCreateUpdateSerializer,MasterListSerializer,
    MasterlistDocumentSerializer,
    DocumentUploadSerializer
)
from django.shortcuts import get_object_or_404

class MasterDropdownView(APIView):
    def get(self, request):
        master_type = request.query_params.get('type')
        if not master_type:
            return Response({"error": "type parameter is required"}, status=400)

        master_type = master_type.upper()

        if master_type == "SUPPLIER":
            data = SupplierSerializer(Supplier.objects.all(), many=True).data
        elif master_type == "GRADE":
            data = GradeSerializer(Grade.objects.all(), many=True).data
        elif master_type == "CUSTOMER":
            data = CustomerSerializer(Customer.objects.all(), many=True).data
        elif master_type == "MATERIAL":
            data = TypeOfMaterialSerializer(TypeOfMaterial.objects.all(), many=True).data
        elif master_type == "LOCATION":
            data = LocationSerializer(Location.objects.all(), many=True).data
        elif master_type == "FORGING_LINE":
            data = LocationSerializer(Location.objects.all(), many=True).data
        else:
            return Response({"error": "Invalid type"}, status=400)

        return Response(data)


# ✅ Efficient pagination
class LargeResultsSetPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = 'page_size'
    max_page_size = 5000


# ------------------- RMReceiving -------------------
class RMReceivingListCreateAPIView(APIView):
    pagination_class = LargeResultsSetPagination
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        try:
            queryset = RMReceiving.objects.all().order_by('-date')

            # 🔍 Filtering logic
            filters = {
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'supplier': request.query_params.get('supplier'),
                'approval_status': request.query_params.get('approval_status'),
                'customer': request.query_params.get('customer'),
                'invoice_no': request.query_params.get('invoice_no'),
                'heatno': request.query_params.get('heatno'),
            }

            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})

            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            if date_from:
                queryset = queryset.filter(date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(date__lte=parse_date(date_to))

            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(queryset, request)
            serializer = RMReceivingSerializer(paginated_qs, many=True)
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            return Response(
                {"error": f"Failed to fetch RMReceiving data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        try:
            # Check if it's bulk create or single create
            is_bulk = isinstance(request.data, list)
            
            with transaction.atomic():
                if is_bulk:
                    serializer = RMReceivingSerializer(data=request.data, many=True)
                else:
                    serializer = RMReceivingSerializer(data=request.data)
                
                serializer.is_valid(raise_exception=True)
                serializer.save()
                
                return Response(
                    {
                        "message": "RMReceiving record(s) created successfully.",
                        "count": len(serializer.data) if is_bulk else 1,
                        "data": serializer.data,
                    },
                    status=status.HTTP_201_CREATED,
                )
        except serializers.ValidationError as e:
            return Response({"error": "Validation error", "details": e.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class RMReceivingDetailAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]  # Add JSONParser

    def get_object(self, uid):
        try:
            return RMReceiving.objects.get(uid=uid)
        except RMReceiving.DoesNotExist:
            raise Http404

    def get(self, request, uid):
        try:
            obj = self.get_object(uid)
            serializer = RMReceivingSerializer(obj)
            return Response(serializer.data)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, uid):
        try:
            obj = self.get_object(uid)
            data = request.data.copy()
            
            # Handle different content types
            if request.content_type == 'application/json':
                # For JSON data, process normally
                serializer = RMReceivingSerializer(obj, data=data, partial=True)
            else:
                # For form data, handle file fields specially
                file_fields = ['milltc', 'spectro', 'ssb_inspection_report', 'customer_approval']
                for field in file_fields:
                    if field in data and data[field] in [None, 'null', '']:
                        del data[field]
                
                # Convert string values for decimal fields
                decimal_fields = ['reciving_weight_kg', 'hold_weight_kg', 'remaining', 'cost_per_kg']
                for field in decimal_fields:
                    if field in data and isinstance(data[field], str):
                        try:
                            data[field] = float(data[field])
                        except (ValueError, TypeError):
                            pass
                
                serializer = RMReceivingSerializer(obj, data=data, partial=True)
            
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            
            print("Serializer errors:", serializer.errors)  # Debug logging
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            print("Exception in patch:", str(e))  # Debug logging
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    
# ------------------- HoldMaterial -------------------
import traceback
import sys
import logging
from django.db.models import Sum, F, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
logger = logging.getLogger(__name__)
class HoldMaterialListCreateAPIView(APIView):
    pagination_class = LargeResultsSetPagination

    def get(self, request):
        try:
            queryset = HoldMaterial.objects.all().order_by('-created_at')

            # 🔍 Collect filters
            filters = {
                'component': request.query_params.get('component'),
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'supplier': request.query_params.get('supplier'),
                'customer': request.query_params.get('customer'),
                'heatno': request.query_params.get('heatno'),
                'rack_no': request.query_params.get('rack_no'),
                'batch_id': request.query_params.get('batch_id'),
            }

            # 🟡 Apply filters dynamically
            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})

            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            if date_from:
                queryset = queryset.filter(created_at__date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(created_at__date__lte=parse_date(date_to))

            # 🔘 Status filter (open / partial / complete / all)
            status_filter = request.query_params.get('status')
            if status_filter and status_filter.lower() != "all":
                queryset = queryset.filter(status__iexact=status_filter)

            # 🧾 Pagination
            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(queryset, request)

            serializer = HoldMaterialListSerializer(paginated_qs, many=True)
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            logger.error("Error fetching HoldMaterial list", exc_info=True)
            return Response(
                {"error": f"Failed to fetch HoldMaterial data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        data = request.data
        is_bulk = isinstance(data, list)

        try:
            with transaction.atomic():
                serializer = HoldMaterialSerializer(data=data, many=is_bulk)
                serializer.is_valid(raise_exception=True)
                instances = serializer.save()

                if not is_bulk:
                    instances = [instances]

                for instance in instances:
                    if instance.rm_receiving:
                        instance.rm_receiving.update_status()

                return Response(
                    {
                        "message": "HoldMaterial record(s) created successfully.",
                        "count": len(serializer.data) if is_bulk else 1,
                        "data": serializer.data,
                    },
                    status=status.HTTP_201_CREATED,
                )

        except Exception as e:
            # Print full traceback in Django terminal
            print("\n🛑 ERROR creating HoldMaterial 🛑")
            print(f"Error Type: {type(e).__name__}")
            print(f"Error Message: {str(e)}")
            traceback.print_exc()

            # Also log it for future reference
            logger.error("HoldMaterial creation failed", exc_info=True)

            # Return the full traceback to frontend (dev only!)
            return Response({
                "error_type": type(e).__name__,
                "error_message": str(e),
                "traceback": traceback.format_exc(),
                "request_data": data
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ------------------- BatchTracking -------------------
class BatchTrackingListCreateAPIView(APIView):
    pagination_class = LargeResultsSetPagination

    def get(self, request):
        try:
            queryset = BatchTracking.objects.select_related('batch_id').order_by('-created_at')

            # 🔍 Filters
            filters = {
                'component': request.query_params.get('component'),
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'customer': request.query_params.get('customer'),
                'heatno': request.query_params.get('heatno'),
                'rack_no': request.query_params.get('rack_no'),
                'issue_id': request.query_params.get('issue_id'),
                'batch_id__batch_id': request.query_params.get('batch_id'),  # ✅ Proper relation filter
            }

            # Apply filters dynamically
            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})
            # 🗓️ Apply date filters
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')

            if date_from:
                queryset = queryset.filter(created_at__date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(created_at__date__lte=parse_date(date_to))
            # Pagination
            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(queryset, request)

            serializer = BatchTrackingSerializer(paginated_qs, many=True)
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            logger.error("Error fetching BatchTracking list", exc_info=True)
            return Response(
                {"error": f"Failed to fetch BatchTracking data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        data = request.data
        is_bulk = isinstance(data, list)

        try:
            with transaction.atomic():
                serializer = BatchTrackingSerializer(data=data, many=is_bulk)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                return Response(
                    {
                        "message": "BatchTracking record(s) created successfully.",
                        "count": len(serializer.data) if is_bulk else 1,
                        "data": serializer.data,
                    },
                    status=status.HTTP_201_CREATED,
                )
        except Exception as e:
            logger.error("BatchTracking creation failed", exc_info=True)
            return Response({
                "error": f"Failed to create BatchTracking: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from .models import Masterlist
from .serializers import MasterlistSerializer,RMReceivingFilterSerializer

class ComponentSuggestionAPIView(APIView):
    def get(self, request):
        query = request.query_params.get("q", "").strip()
        if not query:
            return Response([], status=status.HTTP_200_OK)
        
        # Return top 10 suggestions
        components = (
            Masterlist.objects.filter(component__icontains=query)
            .values_list("component", flat=True)
            .distinct()[:10]
        )
        return Response(components, status=status.HTTP_200_OK)


class ComponentDetailAPIView(APIView):
    def get(self, request):
        component = request.query_params.get("component")
        if not component:
            return Response({"error": "Component required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            obj = Masterlist.objects.get(component=component)
            data = {
                "grade": obj.grade,
                "dia": obj.dia,
                "customer": obj.customer,
                "slug_weight": obj.slug_weight,
                "standerd": obj.standerd,
                "supplier": obj.supplier,
                "location": obj.customer_location,
            }
            return Response(data, status=status.HTTP_200_OK)
        except Masterlist.DoesNotExist:
            return Response({"error": "Component not found"}, status=status.HTTP_404_NOT_FOUND)
        
class RMReceivingFilteredAPIView(APIView):
    """
    Filters RMReceiving by grade, dia, supplier, customer
    Only includes status='open' or 'partial'
    Sorts by remaining (asc) then date (asc)
    Adds f1, f2, ... label in response
    """

    def get(self, request):
        try:
            grade = request.query_params.get("grade")
            dia = request.query_params.get("dia")
            supplier = request.query_params.get("supplier")
            customer = request.query_params.get("customer")

            base_queryset = RMReceiving.objects.filter(status__in=["open", "partial"])

            print(base_queryset)

            # Helper: sequential filter function
            def sequential_filter(field_name, field_values, queryset):
                """Tries filtering by each value in order, returning the first non-empty queryset."""
                for value in field_values:
                    temp_qs = queryset.filter(**{f"{field_name}__icontains": value.strip()})
                    if temp_qs.exists():
                        return temp_qs
                return queryset.none()

            queryset = base_queryset

            # Sequential checks for each multi-value field
            if grade:
                grades = [g.strip() for g in grade.split(",") if g.strip()]
                queryset = sequential_filter("grade", grades, queryset)

            if dia:
                dias = [d.strip() for d in dia.split(",") if d.strip()]
                queryset = sequential_filter("dia", dias, queryset)

            if supplier:
                suppliers = [s.strip() for s in supplier.split(",") if s.strip()]
                queryset = sequential_filter("supplier", suppliers, queryset)

            # Customer is not sequential (usually specific)
            if customer:
                queryset = queryset.filter(customer__icontains=customer)

            queryset = queryset.order_by("remaining", "date")
            print(queryset)

            response_data = []
            for i, obj in enumerate(queryset, start=1):
                serializer = RMReceivingFilterSerializer(obj, context={"index": i})
                data = serializer.data
                data["label"] = f"f{i}"
                response_data.append(data)

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Failed to fetch RMReceiving data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BatchSuggestionAPIView(APIView):
    def get(self, request):
        query = request.query_params.get("q", "").strip()
        if not query:
            return Response([], status=status.HTTP_200_OK)

        batches = (
            HoldMaterial.objects
            .filter(batch_id__icontains=query)
            .values("id", "batch_id")[:10]
        )
        
        # Return data in consistent format
        suggestions = [{"id": b["id"], "batch_id": b["batch_id"]} for b in batches]
        return Response(suggestions, status=status.HTTP_200_OK)

    
class BatchDetailAPIView(APIView):
    def get(self, request):
        batch_id = request.query_params.get("batch_id")
        if not batch_id:
            return Response({"error": "Batch ID required"}, status=400)

        try:
            # Try to get by string batch_id first, then by numeric ID
            if batch_id.isdigit():
                batch = HoldMaterial.objects.get(id=batch_id)
            else:
                batch = HoldMaterial.objects.get(batch_id=batch_id)
        except HoldMaterial.DoesNotExist:
            return Response({"error": "Batch not found"}, status=404)

        serializer = HoldMaterialSerializer(batch)
        return Response(serializer.data)

class IssueBatchSuggestionAPIView(APIView):
    def get(self, request):
        query = request.query_params.get("q", "").strip()
        if not query:
            return Response([], status=status.HTTP_200_OK)

        # Get unique batch_ids from the related Batch model
        batch_ids = (
            BatchTracking.objects
            .filter(batch_id__batch_id__icontains=query)
            .values_list("batch_id__batch_id", flat=True)
            .distinct()
        )

        # Convert to a list and take first 10
        suggestions = [{"batch_id": b} for b in list(batch_ids)[:10]]

        return Response(suggestions, status=status.HTTP_200_OK)
    

def autocomplete_batch(request):
    term = request.GET.get('term', '').strip().lower()
    if term:
        batch_ids = HoldMaterial.objects.filter(block_mt_id__icontains=term).values_list('block_mt_id', flat=True).distinct()
        return JsonResponse(sorted(list(batch_ids)), safe=False)
    return JsonResponse([], safe=False)

# views.py
def batch_details(request):
    batch_id = request.GET.get('batch_id', '').strip()
    try:
        batch = HoldMaterial.objects.get(batch_id=batch_id)
        total_generated = TagGeneration.objects.filter(
            batch_id=batch_id,
            current_process__iexact="Forging"
        ).aggregate(total=Sum('qty'))['total'] or 0

        max_qty=max(batch.pieces-total_generated,0)
        data = {
            'grade': batch.grade,
            'heat_no': batch.heatno,
            'customer': batch.customer,
            'component': batch.component,
            'slug_weight': batch.slug_weight,
            'max_qty': max_qty
        }
        return JsonResponse(data)
    except HoldMaterial.DoesNotExist:
        return JsonResponse({'error': 'Batch not found'}, status=404)
    


@api_view(['GET'])
def batch_remaining_qty(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0
        
        # Get total already generated in this department
        total_generated = TagGeneration.objects.filter(
            batch_id=batch_id,
            current_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        remaining_qty = max(total_received - total_generated, 0)

        return Response({"remaining_qty": remaining_qty})
    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

# views.py
@api_view(['GET'])
def get_child_components(request):
    parent_component = request.GET.get('parent_component', '').strip()
    if not parent_component:
        return Response({'error': 'Parent component is required'}, status=400)

    try:
        parent = Masterlist.objects.get(component__iexact=parent_component)
        children = parent.child_components.all().values_list('component', flat=True)
        data = list(children)
        # include parent itself in dropdown
        data.insert(0, parent.component)
        return Response(data)
    except Masterlist.DoesNotExist:
        return Response({'error': 'Parent component not found'}, status=404)


@api_view(['GET'])
def batch_remaining_qty_forging(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = Forging.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Sum('production'))['total'] or 0
        

        return Response({"remaining_qty": total_received})
    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

class BatchTrackingFieldsAPIView(APIView):
    def get(self, request):
        """Return all field names and verbose names from the Forging model"""
        fields_info = [
            {
                "name": f.name,
                "label": f.verbose_name.replace("_", " ").title(),
                "type": f.get_internal_type()
            }
            for f in BatchTracking._meta.fields
        ]
        return Response(fields_info)
    
class BatchTrackingExportAPIView(APIView):
    def get(self, request):
        try:
            queryset = BatchTracking.objects.all().order_by('-created_at')

            # 🔍 Filters
            filters = {
                'component': request.query_params.get('component'),
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'customer': request.query_params.get('customer'),
                'heatno': request.query_params.get('heatno'),
                'rack_no': request.query_params.get('rack_no'),
                'issue_id': request.query_params.get('issue_id'),
                'batch_id__batch_id': request.query_params.get('batch_id'),  # ✅ Proper relation filter
            }
            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})

            # 🔹 Date range
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            if date_from:
                queryset = queryset.filter(created_at__date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(created_at__date__lte=parse_date(date_to))
            # ✅ Selected fields (comma separated)
            selected_fields = request.query_params.get('fields')
            if selected_fields:
                selected_fields = [f.strip() for f in selected_fields.split(',')]
            else:
                # default: all fields
                selected_fields = [f.name for f in BatchTracking._meta.fields]

            # ✅ Workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Issue Data"

            # Headers
            ws.append([f.replace('_', ' ').title() for f in selected_fields])

            # Rows
            for obj in queryset:
                row = []
                for field in selected_fields:
                    value = getattr(obj, field, '')

                    # ✅ Handle ImageField safely
                    if hasattr(value, 'name') and isinstance(value, models.fields.files.FieldFile):
                        value = value.name if value.name else ''

                    # ✅ Handle ForeignKey (batch_id -> only batch_id string)
                    elif isinstance(value, models.Model):
                        if hasattr(value, 'batch_id'):
                            value = value.batch_id
                        else:
                            value = str(value)

                    # ✅ Handle date objects
                    elif isinstance(value, datetime.date):
                        value = value.strftime('%Y-%m-%d')

                    # ✅ Ensure primitive type
                    elif not isinstance(value, (str, int, float, Decimal)):
                        value = str(value) if value else ''

                    row.append(value)
                ws.append(row)




            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=Issue_data.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            logger.error("Error exporting Forging data", exc_info=True)
            return Response(
                {"error": f"Failed to export Forging data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class RMReceivingFieldsAPIView(APIView):
    def get(self, request):
        """Return all field names and verbose names from the Forging model"""
        fields_info = [
            {
                "name": f.name,
                "label": f.verbose_name.replace("_", " ").title(),
                "type": f.get_internal_type()
            }
            for f in RMReceiving._meta.fields
        ]
        return Response(fields_info)

class RMReceivingExportAPIView(APIView):
    def get(self, request):
        try:
            queryset = RMReceiving.objects.all().order_by('-date')

            # 🔍 Filters
            filters = {
                'component': request.query_params.get('component'),
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'customer': request.query_params.get('customer'),
                'heatno': request.query_params.get('heatno'),
                'rack_no': request.query_params.get('rack_no'),
                'issue_id': request.query_params.get('issue_id'),
                'batch_id__batch_id': request.query_params.get('batch_id'),  # ✅ Proper relation filter
            }
            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})

            # 🔹 Date range
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            if date_from:
                queryset = queryset.filter(date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(date__lte=parse_date(date_to))
            # ✅ Selected fields (comma separated)
            selected_fields = request.query_params.get('fields')
            if selected_fields:
                selected_fields = [f.strip() for f in selected_fields.split(',')]
            else:
                # default: all fields
                selected_fields = [f.name for f in RMReceiving._meta.fields]

            # ✅ Workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RM Reciving Data"

            # Headers
            ws.append([f.replace('_', ' ').title() for f in selected_fields])

            # Rows
            
                        
            for obj in queryset:
                row = []
                for field in selected_fields:
                    value = getattr(obj, field, '')

                    # 🔹 Convert unsupported types
                    if isinstance(value, datetime.date):
                        value = value.strftime('%Y-%m-%d')

                    elif isinstance(value, uuid.UUID):
                        value = str(value)

                    elif isinstance(value, FieldFile):  # ✅ Safely handle FileField / ImageField
                        if value and value.name:
                            try:
                                value = request.build_absolute_uri(value.url)
                            except Exception:
                                value = value.name
                        else:
                            value = ''  # No file uploaded

                    elif hasattr(value, 'id') and hasattr(value, '__str__'):
                        # ✅ Handles ForeignKey or related objects
                        value = str(value)

                    elif isinstance(value, (list, dict, set)):
                        value = str(value)

                    row.append(value)
                ws.append(row)


            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=RM_Reciving.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            logger.error("Error exporting RM data", exc_info=True)
            return Response(
                {"error": f"Failed to export RM data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class HoldMaterialFieldsAPIView(APIView):
    def get(self, request):
        """Return all field names and verbose names from the Forging model"""
        fields_info = [
            {
                "name": f.name,
                "label": f.verbose_name.replace("_", " ").title(),
                "type": f.get_internal_type()
            }
            for f in HoldMaterial._meta.fields
        ]
        return Response(fields_info)

class HoldMaterialExportAPIView(APIView):
    def get(self, request):
        try:
            queryset = HoldMaterial.objects.all().order_by('-created_at')

            # 🔍 Filters
            filters = {
                'component': request.query_params.get('component'),
                'grade': request.query_params.get('grade'),
                'dia': request.query_params.get('dia'),
                'supplier': request.query_params.get('supplier'),
                'customer': request.query_params.get('customer'),
                'heatno': request.query_params.get('heatno'),
                'rack_no': request.query_params.get('rack_no'),
                'batch_id': request.query_params.get('batch_id'),
            }
            for field, value in filters.items():
                if value:
                    queryset = queryset.filter(**{f"{field}__icontains": value})

            # 🔹 Date range
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            if date_from:
                queryset = queryset.filter(created_at__date__gte=parse_date(date_from))
            if date_to:
                queryset = queryset.filter(created_at__date__lte=parse_date(date_to))
            # ✅ Selected fields (comma separated)
            selected_fields = request.query_params.get('fields')
            if selected_fields:
                selected_fields = [f.strip() for f in selected_fields.split(',')]
            else:
                # default: all fields
                selected_fields = [f.name for f in RMReceiving._meta.fields]

            # ✅ Workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RM Hold Data"

            # Headers
            ws.append([f.replace('_', ' ').title() for f in selected_fields])

            # Rows
            
                        
            for obj in queryset:
                row = []
                for field in selected_fields:
                    value = getattr(obj, field, '')

                    # 🔹 Convert unsupported types
                    if isinstance(value, datetime.date):
                        value = value.strftime('%Y-%m-%d')

                    elif isinstance(value, uuid.UUID):
                        value = str(value)

                    elif isinstance(value, FieldFile):  # ✅ Safely handle FileField / ImageField
                        if value and value.name:
                            try:
                                value = request.build_absolute_uri(value.url)
                            except Exception:
                                value = value.name
                        else:
                            value = ''  # No file uploaded

                    elif hasattr(value, 'id') and hasattr(value, '__str__'):
                        # ✅ Handles ForeignKey or related objects
                        value = str(value)

                    elif isinstance(value, (list, dict, set)):
                        value = str(value)

                    row.append(value)
                ws.append(row)


            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=RM_Hold.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            logger.error("Error exporting RM data", exc_info=True)
            return Response(
                {"error": f"Failed to export RM data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        
from .models import TagGeneration
from .serializers import TagGenerationSerializer, TagGenerationCreateSerializer
from rest_framework import viewsets, status
from rest_framework.decorators import action
class TagGenerationViewSet(viewsets.ModelViewSet):
    queryset = TagGeneration.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return TagGenerationCreateSerializer
        return TagGenerationSerializer
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tag = serializer.save()
        # return full details using read serializer
        read_serializer = TagGenerationSerializer(tag, context={'request': request})
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def mark_printed(self, request, pk=None):
        tag = self.get_object()
        tag.is_printed = True
        tag.save()
        return Response({'status': 'marked as printed'})
    
    @action(detail=False, methods=['get'])
    def recent_tags(self, request):
        recent_tags = TagGeneration.objects.all().order_by('-generated_at')[:10]
        serializer = self.get_serializer(recent_tags, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def search_by_uid(self, request):
        tag_uid = request.GET.get('uid', '').strip()
        if not tag_uid:
            return Response(
                {'error': 'Tag UID is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            tag = TagGeneration.objects.get(tag_uid=tag_uid)
            serializer = self.get_serializer(tag)
            return Response(serializer.data)
        except TagGeneration.DoesNotExist:
            return Response(
                {'error': 'Tag not found with the provided UID'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError:
            return Response(
                {'error': 'Invalid UID format'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def search_by_batch_component(self, request):
        batch_id = request.GET.get('batch_id', '').strip()
        component = request.GET.get('component', '').strip()
        
        if not batch_id and not component:
            return Response(
                {'error': 'Please provide either Batch ID or Component'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = TagGeneration.objects.all()
        
        if batch_id:
            queryset = queryset.filter(batch_id__icontains=batch_id)
        if component:
            queryset = queryset.filter(component__icontains=component)
        
        queryset = queryset.order_by('-generated_at')[:20]
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
from heat_treatment.models import HeatTreatment
@api_view(['GET'])
def batch_remaining_qty_heat_treatment(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        prodction_enterd = HeatTreatment.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Sum('production'))['total'] or 0
        
        data = {
            'total_received': total_received,
            'prodction_enterd': prodction_enterd,
            
        }
        return JsonResponse(data)

    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
  

@api_view(['GET'])
def batch_remaining_qty_pre_mc(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        prodction_enterd = pre_mc.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Sum('qty'))['total'] or 0
        
        data = {
            'total_received': total_received,
            'prodction_enterd': prodction_enterd,
            
        }
        return JsonResponse(data)

    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
def batch_remaining_fi(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # 1️⃣ Get total qty received from previous process
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        # 2️⃣ Get total production entered for that batch (setup=II)
        total_production = Fi.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Coalesce(Sum('production'), Value(0)))['total']

        # 3️⃣ Get total rejection counts (sum of all defect fields)
        total_rejections = Fi.objects.filter(
            batch_number=batch_id
        ).aggregate(
            total=Coalesce(
                Sum(
                    F('cnc_height') + F('cnc_od') + F('cnc_bore') + F('cnc_groove') +
                    F('cnc_dent') + F('forging_height') + F('forging_od') +
                    F('forging_bore') + F('forging_crack') + F('forging_dent') +F('rust') +
                    F('pre_mc_height') + F('pre_mc_od') + F('pre_mc_bore')
                ),
                Value(0)
            )
        )['total']

        # 4️⃣ Combine them
        total_entered = total_production + total_rejections

        data = {
            'total_received': total_received,
            'total_production': total_production,
            'total_rejections': total_rejections,
            'total_entered': total_entered,
            'remaining': total_received - total_entered if total_received else 0
        }

        return JsonResponse(data)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


@api_view(['GET'])
def batch_remaining_machining(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # 1️⃣ Get total qty received from previous process
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        # 2️⃣ Get total production entered for that batch (setup=II)
        total_production = machining.objects.filter(
            batch_number=batch_id,
            setup__iexact='II'
        ).aggregate(total=Coalesce(Sum('production'), Value(0)))['total']

        # 3️⃣ Get total rejection counts (sum of all defect fields)
        total_rejections = machining.objects.filter(
            batch_number=batch_id
        ).aggregate(
            total=Coalesce(
                Sum(
                    F('cnc_height') + F('cnc_od') + F('cnc_bore') + F('cnc_groove') +
                    F('cnc_dent') + F('forging_height') + F('forging_od') +
                    F('forging_bore') + F('forging_crack') + F('forging_dent') +
                    F('pre_mc_height') + F('pre_mc_od') + F('pre_mc_bore')
                ),
                Value(0)
            )
        )['total']

        # 4️⃣ Combine them
        total_entered = total_production + total_rejections

        data = {
            'total_received': total_received,
            'total_production': total_production,
            'total_rejections': total_rejections,
            'total_entered': total_entered,
            'remaining': total_received - total_entered if total_received else 0
        }

        return JsonResponse(data)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def batch_remaining_qty_marking(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        prodction_enterd = marking.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Sum('qty'))['total'] or 0
        
        data = {
            'total_received': total_received,
            'prodction_enterd': prodction_enterd,
            
        }
        return JsonResponse(data)

    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def batch_remaining_qty_visual(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # 1️⃣ Get total qty received from previous process
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        # 2️⃣ Get total production entered for that batch (setup=II)
        total_production = Visual.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Coalesce(Sum('production'), Value(0)))['total']

        # 3️⃣ Get total rejection counts (sum of all defect fields)
        total_rejections = Visual.objects.filter(
            batch_number=batch_id
        ).aggregate(
            total=Coalesce(
                Sum(
                    F('cnc_height') + F('cnc_od') + F('cnc_bore') + F('cnc_groove') +
                    F('cnc_dent') +F('cnc_rust') + F('forging_height') + F('forging_od') +
                    F('forging_bore') + F('forging_crack') + F('forging_dent') +
                    F('pre_mc_height') + F('pre_mc_od') + F('pre_mc_bore')+ F('marking')
                ),
                Value(0)
            )
        )['total']

        # 4️⃣ Combine them
        total_entered = total_production + total_rejections

        data = {
            'total_received': total_received,
            'total_production': total_production,
            'total_rejections': total_rejections,
            'total_entered': total_entered,
            'remaining': total_received - total_entered if total_received else 0
        }

        return JsonResponse(data)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
def batch_remaining_qty_dispatch(request):
    batch_id = request.GET.get('batch_id', '').strip()
    current_dept = request.GET.get('current_department', '').strip()

    if not batch_id or not current_dept:
        return Response(
            {"error": "Batch ID and Current Department are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Get total quantity from previous department movements
        # Assume MaterialMovement model tracks movements
        total_received = TagGeneration.objects.filter(
            batch_id=batch_id,
            next_process__iexact=current_dept
        ).aggregate(total=Sum('qty'))['total'] or 0

        prodction_enterd = dispatch.objects.filter(
            batch_number=batch_id,
        ).aggregate(total=Sum('pices'))['total'] or 0

        remaining = total_received - prodction_enterd

        
        data = {
            'total_received': total_received,
            'prodction_enterd': prodction_enterd,
            "remaining": remaining
        
        }
        return JsonResponse(data)

    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
def get_operation_target(request):
    component = request.GET.get('component', '').strip().lower()
    setup = request.GET.get('setup', '').strip().upper()  # Setup should be 'I' or 'II'
    
    # Fetch the masterlist entry for the given component
    if component.endswith('-npd'):
        component = component[:-4]

    try:
        masterlist_entry = Masterlist.objects.get(component__iexact=component)
    except Masterlist.DoesNotExist:
        return JsonResponse({'error': 'Component not found'}, status=404)
    
    # Determine the target based on setup type
    if setup == 'I':
        target = masterlist_entry.op_10_target
    elif setup == 'II':
        target = masterlist_entry.op_20_target
    else:
        return JsonResponse({'error': 'Invalid setup value. Use I or II.'}, status=400)
    
    # Prepare response data
    data = {
        'component': masterlist_entry.component,
        'customer': masterlist_entry.customer,
        'drawing_number': masterlist_entry.drawing_sr_number,
        'setup': setup,
        'target': target if target is not None else 'No target available'
    }
    
    return JsonResponse(data)


class MasterlistAPIView(generics.ListAPIView):
    serializer_class = MasterlistSerializer
    queryset = Masterlist.objects.all()


@api_view(['GET'])
def invoice_list(request):
    # Get filter parameters
    invoice_no = request.GET.get('invoice_no', '')
    heat_no = request.GET.get('heat_no', '')
    supplier = request.GET.get('supplier', '')
    customer = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    material_type = request.GET.get('material_type', '')
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 20)

    # Base query
    query = Q()
    if invoice_no:
        query &= Q(invoice_no__icontains=invoice_no)
    if heat_no:
        query &= Q(heatno__icontains=heat_no)
    if supplier:
        query &= Q(supplier__icontains=supplier)
    if customer:
        query &= Q(customer__icontains=customer)
    if date_from:
        query &= Q(date__gte=date_from)
    if date_to:
        query &= Q(date__lte=date_to)
    if material_type:
        query &= Q(type_of_material__icontains=material_type)

    # Get all unique invoices with summary data
    invoices = RMReceiving.objects.filter(query).values('invoice_no').annotate(
        total_weight=Sum('reciving_weight_kg'),
        heat_count=Count('heatno', distinct=True),
        first_date=Min('date'),
        last_date=Max('date'),
        material_type=Max('type_of_material'),
        total_cost=Sum('reciving_weight_kg') * Sum('cost_per_kg'),
        heat_numbers=GroupConcat('heatno')
    ).order_by('-first_date')

    # Pagination
    paginator = Paginator(invoices, per_page)
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)

    # Get supplier and customer for each invoice
    invoice_data = []
    for inv in page_obj.object_list:
        first_item = RMReceiving.objects.filter(invoice_no=inv['invoice_no']).first()
        invoice_data.append({
            'invoice_no': inv['invoice_no'],
            'supplier': first_item.supplier if first_item else '',
            'customer': first_item.customer if first_item else '',
            'total_weight': inv['total_weight'],
            'heat_count': inv['heat_count'],
            'heat_numbers': inv['heat_numbers'],
            'date_range': f"{inv['first_date']} to {inv['last_date']}",
            'first_date': inv['first_date'],
            'last_date': inv['last_date'],
            'material_type': inv['material_type'],
            'total_cost': inv['total_cost'] if first_item and first_item.cost_per_kg else None
        })

    return Response({
        'invoices': invoice_data,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'total_items': paginator.count
    }, status=status.HTTP_200_OK)
from django.db.models.functions import Lower
@api_view(['GET'])
def invoice_details(request):
    try:
        invoice_no = request.GET.get('invoice_no')
       
        raw_materials = RMReceiving.objects.filter(invoice_no=invoice_no)
        
        if not raw_materials.exists():
            return Response(
                {"error": f"No records found for invoice {invoice_no}"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get all unique heat numbers from the raw materials
        heat_numbers = list(set(rm.heatno for rm in raw_materials))
        raw_materials = RMReceiving.objects.filter(heatno__in=heat_numbers)
        
        # Calculate total raw material weight
        total_rm_weight = sum(float(rm.reciving_weight_kg) for rm in raw_materials)
        
        # Prepare response data
        response_data = {
            "invoice_no": invoice_no,
            "supplier": raw_materials[0].supplier,
            "customer": raw_materials[0].customer,
            "material_type": raw_materials[0].type_of_material,
            "heat_numbers": heat_numbers,
            "raw_materials": [],
            "block_materials": [],
            "dispatch_details": [],
            "production_details": [],
            "cost_summary": {
                "total_raw_material_cost": 0,
                "total_production_value": 0,
                "total_dispatch_value": 0
            },
            "yield_summary": {
                "total_rm_weight": total_rm_weight,
                "total_dispatch_weight": 0,
                "total_slug_weight": 0,
                "yield_percentage": 0
            }
        }
        
        # Raw Material Details with cost calculation
        for rm in raw_materials:
            rm_cost = float(rm.reciving_weight_kg) * float(rm.cost_per_kg) if rm.cost_per_kg else 0
            response_data["raw_materials"].append({
                "date": rm.date,
                "supplier": rm.supplier,
                "grade": rm.grade,
                "customer": rm.customer,
                "heatno": rm.heatno,
                "dia": rm.dia,
                "weight": float(rm.reciving_weight_kg),
                "invoice_no": rm.invoice_no,
                "type_of_material": rm.type_of_material,
                "rack_no": rm.rack_no,
                "location": rm.location,
                "cost_per_kg": float(rm.cost_per_kg) if rm.cost_per_kg else None,
                "total_cost": rm_cost,
                "is_job_work": not bool(rm.cost_per_kg)  # Job work if no cost
            })
            response_data["cost_summary"]["total_raw_material_cost"] += rm_cost
        
        # Block Material Details with component cost and completion status
        heat_numbers = list(set(str(rm.heatno).strip().lower() for rm in raw_materials))

        block_materials = HoldMaterial.objects.annotate(
            heatno_lower=Lower('heatno')
        ).filter(heatno_lower__in=heat_numbers)
        print(block_materials)
        print(heat_numbers)
        for bm in block_materials:
            try:
                component_key = bm.component.replace("-NPD", "") if bm.component.endswith("-NPD") else bm.component
                component_cost = Masterlist.objects.get(component=component_key).cost
            except Masterlist.DoesNotExist:
                component_cost = 0
            
            # Calculate remaining quantity and completion status
            total_planned = bm.pieces
            total_dispatched = sum(dd.pices for dd in dispatch.objects.filter(component=bm.component, heat_no=bm.heatno))
            remaining = total_planned - total_dispatched
            is_completed = "Complete" if remaining <= 50 else f"Remaining: {remaining}"
            
            response_data["block_materials"].append({
                "block_mt_id": bm.batch_id,
                "component": bm.component,
                "customer": bm.customer,
                "supplier": bm.supplier,
                "grade": bm.grade,
                "heatno": bm.heatno,
                "dia": bm.dia,
                "pices": bm.pieces,
                "weight": float(bm.hold_material_qty_kg),
                "created_at": bm.created_at,
                "standerd": bm.standerd,
                "line": bm.line,
                "component_cost": component_cost,
                "total_cost": bm.pieces * component_cost,
                "completion_status": is_completed,
                "total_planned": total_planned,
                "total_dispatched": total_dispatched,
                "remaining": remaining
            })
        
        # Production Details from Forging model (to get slug weight)
        forging_details = Forging.objects.filter(heat_number__in=heat_numbers)
        total_slug_weight = 0
        for fd in forging_details:
            try:
                component_key = fd.component.replace("-NPD", "") if fd.component.endswith("-NPD") else fd.component
                component_cost = Masterlist.objects.get(component=component_key).cost

                slug_weight = float(fd.slug_weight) if fd.slug_weight else 0
                total_slug_weight += slug_weight
            except Masterlist.DoesNotExist:
                component_cost = 0
                slug_weight = 0
            
            production_value = fd.production * component_cost
            response_data["production_details"].append({
                "batch_number": fd.batch_number,
                "date": fd.date,
                "shift": fd.shift,
                "component": fd.component,
                "customer": fd.customer,
                "heat_number": fd.heat_number,
                "line": fd.line,
                "target": fd.target,
                "production": fd.production,
                "rework": fd.rework,
                "slug_weight": slug_weight,
                "component_cost": component_cost,
                "production_value": production_value
            })
            response_data["cost_summary"]["total_production_value"] += production_value
        
        # Dispatch Details with value calculation and weight (using slug weight from Forging)
        # Dispatch Details with value calculation and weight (check Forging -> Masterlist -> 0)
        dispatch_details = dispatch.objects.filter(heat_no__in=heat_numbers)
        total_dispatch_weight = 0

        for dd in dispatch_details:
            try:
                component_key = dd.component.replace("-NPD", "") if dd.component.endswith("-NPD") else dd.component
                component_cost = Masterlist.objects.get(component=component_key).cost

            except Masterlist.DoesNotExist:
                component_cost = 0

            # Initialize slug weight
            component_weight = 0.0

            # Check Forging first
            forging_entry = Forging.objects.filter(component=dd.component, heat_number=dd.heat_no).first()
            if forging_entry and forging_entry.slug_weight and float(forging_entry.slug_weight) > 0:
                component_weight = float(forging_entry.slug_weight)
            else:
                # Check Masterlist if Forging doesn't have a valid slug_weight
                try:
                    master_entry = Masterlist.objects.get(component=dd.component)
                    if master_entry.slug_weight and float(master_entry.slug_weight) > 0:
                        component_weight = float(master_entry.slug_weight)
                except Masterlist.DoesNotExist:
                    component_weight = 0.0  # fallback to 0 if not found

            dispatch_weight = dd.pices * component_weight
            total_dispatch_weight += dispatch_weight
            dispatch_value = dd.pices * component_cost

            response_data["dispatch_details"].append({
                "date": dd.date,
                "component": dd.component,
                "pices": dd.pices,
                "invoiceno": dd.invoiceno,
                "heat_no": dd.heat_no,
                "batch_number": dd.batch_number,
                "target1": dd.target1,
                "total_produced": dd.total_produced,
                "remaining": dd.remaining,
                "component_cost": component_cost,
                "slug_weight": component_weight,
                "dispatch_weight": dispatch_weight,
                "total_weight": dd.pices * component_weight,
                "dispatch_value": dispatch_value
            })
            response_data["cost_summary"]["total_dispatch_value"] += dispatch_value

        
        # Update yield summary
        response_data["yield_summary"]["total_dispatch_weight"] = total_dispatch_weight
        response_data["yield_summary"]["total_slug_weight"] = total_slug_weight
        if total_rm_weight > 0:
            response_data["yield_summary"]["yield_percentage"] = (total_dispatch_weight / total_rm_weight) * 100
        
        # Add summary information
        response_data["summary"] = {
            "total_received": sum(rm['weight'] for rm in response_data["raw_materials"]),
            "total_blocks_created": len(response_data["block_materials"]),
            "total_dispatch_records": len(response_data["dispatch_details"]),
            "unique_components": len(set(bm['component'] for bm in response_data["block_materials"])),
            "unique_customers": len(set(bm['customer'] for bm in response_data["block_materials"]))
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

from .models import Schedule
from .serializers import ScheduleSerializer,ScheduleUpdateSerializer
from datetime import datetime

class ScheduleAPIView1(APIView):
    """
    Enhanced API endpoint for schedule management with:
    - Date-based filtering (single date, month, range)
    - Schedule creation (POST)
    - Schedule updates (PUT)
    - Duplicate checking
    """
    permission_classes = []

    def get(self, request):
        """
        Get schedules with enhanced duplicate checking
        """
        date_param = request.query_params.get('date', None)
        component = request.query_params.get('component', None)
        exact_date = request.query_params.get('exact_date', None)
        check_duplicates = request.query_params.get('check_duplicates', 'false').lower() == 'true'
        
        # Enhanced duplicate checking - return all duplicates for the month
        if component and exact_date and check_duplicates:
            try:
                parsed_date = datetime.strptime(exact_date, '%Y-%m-%d').date()
                month_str = f"{parsed_date.year}-{parsed_date.month:02d}"
                
                # Get all existing schedules in the same month
                existing = Schedule.objects.filter(
                    component=component,
                    date1__contains=month_str
                ).order_by('date1')
                
                if existing.exists():
                    return Response({
                        "exists": True,
                        "schedules": ScheduleSerializer(existing, many=True).data,
                        "message": f"This component already has {existing.count()} schedule(s) for {month_str}"
                    })
                return Response({"exists": False})
            except Exception as e:
                logger.error(f"Error checking existing schedule: {str(e)}")
                return Response(
                    {"error": "Failed to check existing schedule"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        # Enhanced duplicate checking
        if component and exact_date:
            try:
                # Parse date to match your date1 format (YYYY-MM-DD)
                parsed_date = datetime.strptime(exact_date, '%Y-%m-%d').date()
                month_str = f"{parsed_date.year}-{parsed_date.month:02d}"
                
                # Check for existing schedule in the same month
                existing = Schedule.objects.filter(
                    component=component,
                    date1__contains=month_str
                ).first()
                
                if existing:
                    return Response({
                        "exists": True,
                        "schedule": ScheduleSerializer(existing).data,
                        "message": f"This component already has a schedule for {month_str}"
                    })
                return Response({"exists": False})
            except Exception as e:
                logger.error(f"Error checking existing schedule: {str(e)}")
                return Response(
                    {"error": "Failed to check existing schedule"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
        if not date_param:
            return Response(
                {"error": "Date parameter is required. Use 'date' query parameter."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Check for date range (format: YYYY-MM-DD:YYYY-MM-DD)
            if ':' in date_param:
                start_date_str, end_date_str = date_param.split(':')
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                
                if start_date > end_date:
                    return Response(
                        {"error": "Start date must be before end date"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                queryset = Schedule.objects.filter(
                    Q(date1__icontains=start_date_str) | 
                    Q(date1__icontains=end_date_str) |
                    Q(date1__range=(start_date_str, end_date_str))
                ).order_by('-created_at')
            
            # Check for month (format: YYYY-MM)
            elif len(date_param.split('-')) == 2:
                year, month = date_param.split('-')
                queryset = Schedule.objects.filter(
                    date1__icontains=f"{year}-{month.zfill(2)}-"
                ).order_by('-created_at')
            
            # Single date (format: YYYY-MM-DD)
            else:
                queryset = Schedule.objects.filter(
                    date1__icontains=date_param
                ).order_by('-created_at')
            
            queryset = queryset.only(
                'id', 'component', 'customer', 'supplier', 'grade', 
                'standerd', 'dia', 'slug_weight', 'pices', 
                'weight', 'date1', 'location', 'verified_by', 'created_at'
            )
            
            serializer = ScheduleSerializer(queryset, many=True)
            return Response(serializer.data)
        
        except ValueError as e:
            logger.error(f"Invalid date format: {str(e)}")
            return Response(
                {"error": f"Invalid date format. {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Unexpected error in GET: {str(e)}")
            return Response(
                {"error": "An unexpected error occurred"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        """
        Create new schedule with enhanced duplicate checking and force-create option
        """
        data = request.data.copy()
        component = data.get('component')
        date_str = data.get('date1')
        existing = None

        # Parse date and check for existing entry
        if component and date_str:
            try:
                parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                month_str = f"{parsed_date.year}-{parsed_date.month:02d}"
                
                existing = Schedule.objects.filter(
                    component=component,
                    date1__contains=month_str
                ).first()

            except Exception as e:
                logger.error(f"Date parsing error: {str(e)}")
                return Response({"error": "Invalid date format"}, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Read force_create from either body or query params and normalize
        force_create = data.get('force_create') or request.query_params.get('force_create')
        force_create = str(force_create).lower() == 'true'

        # If duplicate exists and force_create not set, return error
        if existing and not force_create:
            return Response({
                "error": "Duplicate schedule exists",
                "details": ScheduleSerializer(existing).data
            }, status=status.HTTP_409_CONFLICT)

        # Auto-calculate weight if not provided
        if 'weight' not in data and 'pices' in data and 'slug_weight' in data:
            try:
                data['weight'] = float(data['slug_weight']) * int(data['pices'])
            except (ValueError, TypeError) as e:
                logger.warning(f"Failed to calculate weight: {str(e)}")

        serializer = ScheduleSerializer(data=data)

        try:
            if serializer.is_valid():
                # Validate against negative or zero quantities
                if int(data.get('pices', 0)) <= 0:
                    return Response(
                        {"error": "Quantity must be positive"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Set verified_by if not provided
                if 'verified_by' not in data or not data['verified_by']:
                    serializer.validated_data['verified_by'] = request.user.get_full_name() or request.user.username

                serializer.save()
                logger.info(f"Created new schedule: {serializer.data}")
                return Response(serializer.data, status=status.HTTP_201_CREATED)

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            logger.error(f"Error creating schedule: {str(e)}")
            return Response(
                {"error": "Failed to create schedule", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def put(self, request, pk=None):
        """
        Update existing schedule with quantity validation
        """
        try:
            schedule = Schedule.objects.get(pk=pk)
            current_pieces = schedule.pices
            current_weight = schedule.weight
            
            serializer = ScheduleUpdateSerializer(
                schedule, 
                data=request.data, 
                partial=True
            )
            
            if serializer.is_valid():
                new_pieces = serializer.validated_data.get('pices', current_pieces)
                
                # Validate against negative quantities
                if new_pieces <= 0:
                    return Response(
                        {"error": "Quantity must be positive"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # If weight not provided, calculate based on slug_weight
                if 'weight' not in request.data and hasattr(schedule, 'slug_weight'):
                    serializer.validated_data['weight'] = float(schedule.slug_weight) * new_pieces
                
                serializer.save()
                logger.info(f"Updated schedule {pk}: {serializer.data}")
                return Response(serializer.data)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        except Schedule.DoesNotExist:
            logger.error(f"Schedule not found: {pk}")
            return Response(
                {"error": "Schedule not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error updating schedule {pk}: {str(e)}")
            return Response(
                {
                    "error": "Failed to update schedule",
                    "details": str(e),
                    "field_errors": serializer.errors if 'serializer' in locals() else None
                }, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
# In your views.py
class ScheduleUpdatePlannedView(APIView):
    def put(self, request, pk):
        try:
            schedule = Schedule.objects.get(pk=pk)
            planned = request.data.get('planned', 0)
            
            # Update planned quantity (add to existing if any)
            schedule.planned = (schedule.planned or 0) + int(planned)
            schedule.save()
            
            return Response(
                {"message": "Planned quantity updated successfully"},
                status=status.HTTP_200_OK
            )
            
        except Schedule.DoesNotExist:
            return Response(
                {"error": "Schedule not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

from rest_framework import status as drf_status
from itertools import groupby
from operator import attrgetter

class OpenAndPartialRMReceiving(APIView):

    def normalize_rack(self, rack):
        if not rack:
            return None
        rack = str(rack).strip().upper()

        if rack.isdigit():
            return f"R{int(rack)}"

        return rack

    def get(self, request):
        try:
            rows = list(
                RMReceiving.objects.filter(status__in=["open", "partial"])
                .values(
                    "id", "grade", "customer", "supplier", "dia",
                    "date", "remaining", "status", "rack_no",
                    "approval_status", "heatno","reciving_weight_kg"
                )
            )

            # normalize rack numbers
            for r in rows:
                r["rack_no"] = self.normalize_rack(r["rack_no"])

            print("rows count:", len(rows))

            rows = sorted(rows, key=lambda x: (
                (x["grade"] or "").lower(),
                (x["customer"] or "").lower(),
                (x["supplier"] or "").lower(),
                (x["dia"] or "").lower(),
                x["date"]
            ))

            def group_key(x):
                return (
                    (x["grade"] or "").lower(),
                    (x["customer"] or "").lower(),
                    (x["supplier"] or "").lower(),
                    (x["dia"] or "").lower(),
                )

            ALL_RACKS = [f"R{i}" for i in range(1, 78)]
            used_racks = {r["rack_no"] for r in rows if r["rack_no"]}
            missing_racks = [r for r in ALL_RACKS if r not in used_racks]

            response_data = []

            for _, group in groupby(rows, key=group_key):
                group_list = list(group)
                for idx, item in enumerate(group_list, start=1):
                    item["fifo_number"] = f"f{idx}"
                    response_data.append(item)

            print("response count:", len(response_data))

            return Response({
                "rows": response_data,
                "missing_racks": missing_racks,
                "used_racks": list(used_racks)
            }, status=200)

        except Exception as e:
            print("ERROR:", e)
            return Response({"error": str(e)}, status=500)
        
from django.views.decorators.http import require_GET


@require_GET
def production_data_api(request):
    # Get month and year from query parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    if not month or not year:
        return JsonResponse({'error': 'Month and year parameters are required'}, status=400)
    
    try:
        month = int(month)
        year = int(year)
        if month < 1 or month > 12:
            raise ValueError
    except ValueError:
        return JsonResponse({'error': 'Invalid month or year format'}, status=400)
    
    # Prepare response structure
    response_data = {
        'blockmt': [],
        'schedule': [],
        'forging': [],
        'month': month,
        'year': year
    }
    
    # Helper function to get masterlist data for a component
    def get_masterlist_data(component):
        try:
            master = Masterlist.objects.filter(component=component).first()
            if master:
                return {
                    'customer': master.customer,
                    'location': master.customer_location,
                    'cost': str(master.cost) if master.cost else None
                }
            return None
        except Exception as e:
            print(f"Error fetching masterlist for {component}: {str(e)}")
            return None
    
    # Fetch Blockmt data
    try:
        block_mt_prefix = f"PP-{year}{str(month).zfill(2)}"
        blockmt_qs = HoldMaterial.objects.filter(
            
            batch_id__startswith=block_mt_prefix
        )
        print(blockmt_qs)
        
        for item in blockmt_qs:
            master_data = get_masterlist_data(item.component)
            response_data['blockmt'].append({
                'block_mt_id': item.batch_id,
                'component': item.component,
                'customer': item.customer if item.customer else (master_data['customer'] if master_data else None),
                'supplier': item.supplier,
                'grade': item.grade,
                'standerd': item.standerd,
                'heatno': item.heatno,
                'dia': item.dia,
                'rack_no': item.rack_no,
                'pices': item.pieces,
                'weight': str(item.hold_material_qty_kg),
                'line': item.line,
                'created_at': item.created_at.isoformat(),
                'verified_by': item.verified_by,
                'location': master_data['location'] if master_data else None,
                'cost': master_data['cost'] if master_data else None
            })
    except Exception as e:
        print(f"Error processing Blockmt data: {str(e)}")
    
    # Fetch Schedule data
    try:
        # Since Schedule uses date1 as string, we need to handle it differently
        # Assuming date1 is in format 'YYYY-MM-DD' or similar
        date_prefix = f"{year}-{month:02d}"
        schedule_qs = Schedule.objects.filter(date1__startswith=date_prefix)
        
        for item in schedule_qs:
            master_data = get_masterlist_data(item.component)
            response_data['schedule'].append({
                'component': item.component,
                'customer': item.customer if item.customer else (master_data['customer'] if master_data else None),
                'supplier': item.supplier,
                'grade': item.grade,
                'standerd': item.standerd,
                'dia': item.dia,
                'slug_weight': str(item.slug_weight),
                'pices': item.pices,
                'weight': str(item.weight),
                'date1': item.date1,
                'location': item.location if item.location else (master_data['location'] if master_data else None),
                'verified_by': item.verified_by,
                'created_at': item.created_at.isoformat(),
                'cost': master_data['cost'] if master_data else None
            })
    except Exception as e:
        print(f"Error processing Schedule data: {str(e)}")
    
    # Fetch Forging data
    try:
        forging_qs = Forging.objects.filter(
            date__month=month,
            date__year=year
        )
        
        for item in forging_qs:
            master_data = get_masterlist_data(item.component)
            response_data['forging'].append({
                'batch_number': item.batch_number,
                'date': item.date.isoformat(),
                'shift': item.shift,
                'component': item.component,
                'customer': item.customer if item.customer else (master_data['customer'] if master_data else None),
                'slug_weight': str(item.slug_weight),
                'rm_grade': item.rm_grade,
                'heat_number': item.heat_number,
                'line': item.line,
                'line_incharge': item.line_incharge,
                'forman': item.forman,
                'target': item.target,
                'production': item.production,
                'rework': item.rework,
                'up_setting': item.up_setting,
                'half_piercing': item.half_piercing,
                'full_piercing': item.full_piercing,
                'ring_rolling': item.ring_rolling,
                'sizing': item.sizing,
                'overheat': item.overheat,
                'bar_crack_pcs': item.bar_crack_pcs,
                'verified_by': item.verified_by,
                'machine_status': item.machine_status,
                'downtime_minutes': item.downtime_minutes,
                'reason_for_downtime': item.reason_for_downtime,
                'reason_for_low_production': item.reason_for_low_production,

                'location': master_data['customer_location'] if master_data else None,
                'cost': master_data['cost'] if master_data else None
            })
    except Exception as e:
        print(f"Error processing Forging data: {str(e)}")
    
    return JsonResponse(response_data)



@api_view(['GET', 'POST'])
@permission_classes([])
def masterlist_list_create(request):
    if request.method == 'GET':
        # Pagination - use offset/limit instead of page
        offset = int(request.query_params.get('offset', 0))
        limit = int(request.query_params.get('limit', 20))
        
        # Initialize queryset
        queryset = Masterlist.objects.all().order_by('-created_at')
        
        # Search filter
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(component__icontains=search) |
                Q(part_name__icontains=search) |
                Q(drawing_sr_number__icontains=search)
            )
        
        # Additional filters
        component = request.query_params.get('component')
        if component:
            queryset = queryset.filter(component__icontains=component)

        running_status = request.query_params.get('running_status')
        if running_status:
            queryset = queryset.filter(running_status=running_status)

        customer = request.query_params.get('customer')
        if customer:
            queryset = queryset.filter(customer__icontains=customer)
            
        material_grade = request.query_params.get('material_grade')
        if material_grade:
            queryset = queryset.filter(grade__icontains=material_grade)
            
        ht_process = request.query_params.get('ht_process')
        if ht_process:
            queryset = queryset.filter(ht_process__iexact=ht_process)
        
        # Get distinct values for filter options
        customers = Masterlist.objects.order_by().values_list('customer', flat=True).distinct()
        materials = Masterlist.objects.order_by().values_list('grade', flat=True).distinct()
        ht_processes = Masterlist.objects.order_by().values_list('ht_process', flat=True).distinct()
        
        # Use select_related/prefetch_related for related data
        queryset = queryset.prefetch_related('documents')

        # Get total count before pagination
        total_count = queryset.count()
        
        # Apply pagination using offset/limit
        queryset = queryset[offset:offset + limit]
        
        serializer = MasterlistSerializer1(queryset, many=True, context={'request': request})
        return Response({
            'results': serializer.data,
            'count': total_count,
            'filter_options': {
                'customers': list(customers),
                'materials': list(materials),
                'ht_processes': list(ht_processes)
            }
        })
    
    elif request.method == 'POST':
        serializer = MasterlistCreateUpdateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=status.HTTP_201_CREATED)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([])
def masterlist_retrieve_update_delete(request, pk):
    masterlist = get_object_or_404(Masterlist, pk=pk)
    
    if request.method == 'GET':
        serializer = MasterlistSerializer1(masterlist, context={'request': request})
        return JsonResponse(serializer.data)
    
    elif request.method == 'PUT':
        serializer = MasterlistCreateUpdateSerializer(masterlist, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        masterlist.delete()
        return JsonResponse({'message': 'Masterlist deleted successfully'}, status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
@permission_classes([])
def masterlist_history(request, pk):
    masterlist = get_object_or_404(Masterlist, pk=pk)
    history = masterlist.history.all()
    data = [{
        'history_date': h.history_date,
        'history_user': h.history_user.username if h.history_user else None,
        'changes': [{
            'field': change.field,
            'old': change.old,
            'new': change.new
        } for change in h.diff_against(h.prev_record).changes] if h.prev_record else [],
        'version': h.history_id
    } for h in history]
    return JsonResponse(data, safe=False)

@api_view(['GET'])
@permission_classes([])
def document_list(request, masterlist_pk):
    documents = MasterlistDocument.objects.filter(masterlist_id=masterlist_pk)
    serializer = MasterlistDocumentSerializer(documents, many=True, context={'request': request})
    return JsonResponse(serializer.data, safe=False)

@api_view(['POST'])
@permission_classes([])
def document_upload(request, masterlist_pk):
    masterlist = get_object_or_404(Masterlist, pk=masterlist_pk)
    serializer = DocumentUploadSerializer(data=request.data)
    print("Incoming request data:", request.data)
    
    if serializer.is_valid():
        # Mark old version as not current if exists
        MasterlistDocument.objects.filter(
            masterlist=masterlist,
            document_type=serializer.validated_data['document_type'],
            is_current=True
        ).update(is_current=False)
        
        # Get next version number
        last_version = MasterlistDocument.objects.filter(
            masterlist=masterlist,
            document_type=serializer.validated_data['document_type']
        ).order_by('-version').first()
        
        new_version = last_version.version + 1 if last_version else 1
        
        # Create new document
        document = MasterlistDocument.objects.create(
            masterlist=masterlist,
            document_type=serializer.validated_data['document_type'],
            document=serializer.validated_data['document'],
            version=new_version,
            is_current=True,
            remarks=serializer.validated_data.get('remarks'),
            verified_by=serializer.validated_data.get('verified_by')
        )
        
        return JsonResponse(
            MasterlistDocumentSerializer(document, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )
    
    return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([])
def document_type_history(request, masterlist_pk, doc_type):
    documents = MasterlistDocument.objects.filter(
        masterlist_id=masterlist_pk,
        document_type=doc_type
    ).order_by('-version')
    serializer = MasterlistDocumentSerializer(documents, many=True, context={'request': request})
    return JsonResponse(serializer.data, safe=False)

@api_view(['POST'])
@permission_classes([])
def document_set_current(request, masterlist_pk, doc_pk):
    document = get_object_or_404(MasterlistDocument, pk=doc_pk, masterlist_id=masterlist_pk)
    
    # Mark all documents of this type as not current
    MasterlistDocument.objects.filter(
        masterlist_id=masterlist_pk,
        document_type=document.document_type,
        is_current=True
    ).update(is_current=False)
    
    # Set this document as current
    document.is_current = True
    document.save()
    
    return JsonResponse(
        MasterlistDocumentSerializer(document, context={'request': request}).data
    )


@api_view(['GET'])
@permission_classes([])
def missing_documents_report(request):
    # Expected document types (match your frontend constant)
    EXPECTED_DOCUMENT_TYPES = [
        'Design Records',
        'Authorized Engineering Change Documents',
        'Customer Engineering Approval',
        'Design Failure Modes and Effects Analysis (DFMEA)',
        'Process Flow Diagram',
        'Process Failure Modes and Effects Analysis (PFMEA)',
        'Control Plan',
        'Measurement Systems Analysis (MSA)',
        'Dimensional Results',
        'Records of Material & Performance Test Results',
        'Initial Process Studies',
        'Qualified Laboratory Documentation',
        'Appearance Approval Report (AAR)',
        'Sample Production Parts',
        'Master Sample',
        'Checking Aids',
        'Customer-Specific Requirements',
        'Part Submission Warrant (PSW)',
    ]
    
    # Get all masterlist items with their documents
    masterlists = Masterlist.objects.prefetch_related('documents').filter(
        running_status="Running"
    )

    
    report = []
    
    for masterlist in masterlists:
        # Get all document types that exist for this component
        existing_doc_types = set(masterlist.documents.values_list('document_type', flat=True))
        
        # Find missing document types
        missing_docs = [doc_type for doc_type in EXPECTED_DOCUMENT_TYPES if doc_type not in existing_doc_types]
        
        if missing_docs:
            report.append({
                'component_id': masterlist.id,
                'component_name': masterlist.component,
                'part_name': masterlist.part_name,
                'drawing_number': masterlist.drawing_sr_number,
                'customer': masterlist.customer,
                'missing_documents': missing_docs,
                'missing_count': len(missing_docs)
            })
    
    # Sort by components with most missing documents first
    report.sort(key=lambda x: x['missing_count'], reverse=True)
    
    return JsonResponse({'report': report})


class MasterlistCreateAPIView(generics.CreateAPIView):
    queryset = Masterlist.objects.all()
    serializer_class = MasterListSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        if not serializer.is_valid():
            print("\n❌ MASTERLIST CREATE ERROR:")
            print(serializer.errors)       # <-- this will show exact field error
            return Response(
                {"status": "error", "errors": serializer.errors},
                status=400
            )

        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            {"status": "success", "data": serializer.data},
            status=201,
            headers=headers
        )

from .models import SPCDimension, SPCRecord
from rest_framework import status

class ComponentSPCDetailView(APIView):
    def get(self, request, component):
        dimensions = SPCDimension.objects.filter(component=component).prefetch_related('records')
        response_data = []

        for dim in dimensions:
            spc_records = list(dim.records.all())  # FK reverse accessor

            latest_record = spc_records[0] if spc_records else None
            previous_records = spc_records[1:] if len(spc_records) > 1 else []

            response_data.append({
                'dimension': dim.dimension,
                'name': dim.name,
                'type': dim.type,
                'instrument': dim.instrument,
                'remark': dim.remark,
                'spc_time_period_days': dim.spc_time_period_days,
                'latest_record': {
                    'cp_value': latest_record.cp_value,
                    'cpk_value': latest_record.cpk_value,
                    'uploaded_at': latest_record.uploaded_at,
                    'uploaded_by': latest_record.uploaded_by,
                    'spc_file': latest_record.spc_file.url,
                } if latest_record else None,
                'previous_records': [
                    {
                        'cp_value': record.cp_value,
                        'cpk_value': record.cpk_value,
                        'uploaded_at': record.uploaded_at,
                        'uploaded_by': record.uploaded_by,
                        'spc_file': record.spc_file.url,
                    }
                    for record in previous_records
                ]
            })

        return Response(response_data, status=status.HTTP_200_OK)
    
from .serializers import SPCDimensionSerializer

class BulkSPCDimensionCreateAPIView(APIView):
    def post(self, request):
        # Simply use the data as sent from the frontend
        dimensions_data = request.data.get('dimensions', [])
        
        serializer = SPCDimensionSerializer(data=dimensions_data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    